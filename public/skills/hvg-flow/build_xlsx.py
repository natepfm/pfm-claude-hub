#!/usr/bin/env python3
"""
HVG flow — build the project audit spreadsheet.

Usage:
    python3 build_xlsx.py <config.json> <output.xlsx>

The config JSON shape:
{
  "project": {
    "name":      "<canonical Notion project name>",
    "slug":      "<derived slug used in filenames>",
    "vertical":  "Auto - Forms",
    "imgName":   "<shared ref filename if Mode A; otherwise '(per-line — see Prompts sheet)'>",
    "createdAt": "2026-05-07",
    "user":      "Sam",
    "mb":        "David V",
    "notionUrl": "https://www.notion.so/...",
    "notes":     "optional context line shown on Summary sheet"
  },
  "generation": {
    "modelDisplay":     "Veo 3.1 Fast",
    "mcpModel":         "veo3_1",
    "mcpParams":        {"model": "veo-3-1-fast"},
    "duration":         8,
    "aspectRatio":      "16:9",
    "countPerPrompt":   1,
    "estCostPerClip":   25
  },
  "reference": {
    "mode":             "A" | "B" | "C" | "D" | "E",
    "modeDescription":  "Single shared" | "Per-line" | "Rotating pool" | "Start + end frames" | "Mixed",
    "rotationStrategy": "" | "random" | "sequential" | "script-tagged"   // only for C/E
  },
  "prompts": [
    {
      "lineNum":    1,
      "slug":       "<slug>_L01",
      "state":      "",
      "slide":      "<slide directive from Notion, optional>",
      "dialogue":   "<the line>",
      "references": "<display string: filename, or 'start.png → end.png' for Mode D>",
      "fullPrompt": "<JSON master with id/title/dialogue injected, as a string>",
      "status":     "pending" | "✓" | "Partial" | "✗" | "skipped",
      "v01":        "<filename or empty>",
      "v02":        "<filename or empty>",
      "notes":      ""
    }
  ]
}

Status values render with color:
  ✓        → green
  Partial  → yellow
  ✗        → red
  pending  → grey
  skipped  → grey (italics)
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# === Style palette ===
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_VALUE_FILL = PatternFill("solid", fgColor="DDEBF7")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE"); GOOD_FONT = Font(color="006100", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFEB9C"); WARN_FONT = Font(color="9C5700", bold=True)
BAD_FILL  = PatternFill("solid", fgColor="FFC7CE"); BAD_FONT  = Font(color="9C0006", bold=True)
PEND_FILL = PatternFill("solid", fgColor="E7E6E6"); PEND_FONT = Font(color="595959", bold=True)
SKIP_FILL = PatternFill("solid", fgColor="E7E6E6"); SKIP_FONT = Font(color="595959", italic=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_STYLE = {
    "✓":       (GOOD_FILL, GOOD_FONT),
    "Partial": (WARN_FILL, WARN_FONT),
    "✗":       (BAD_FILL,  BAD_FONT),
    "pending": (PEND_FILL, PEND_FONT),
    "skipped": (SKIP_FILL, SKIP_FONT),
}


def build(config: dict, output_path: str) -> dict:
    project = config["project"]
    generation = config["generation"]
    reference = config.get("reference", {"mode": "A", "modeDescription": "Single shared", "rotationStrategy": ""})
    prompts_data = config["prompts"]

    wb = Workbook()

    # === Summary sheet ===
    summary = wb.active
    summary.title = "Summary"

    summary.merge_cells("A1:B1")
    title_cell = summary["A1"]
    title_cell.value = "HVG Flow — Project Audit"
    title_cell.fill = HEADER_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 28

    total_clips = len(prompts_data) * generation["countPerPrompt"]
    delivered = sum(1 for p in prompts_data if p["status"] == "✓") * generation["countPerPrompt"]
    delivered += sum(1 for p in prompts_data if p["status"] == "Partial")  # 1 of 2

    cost_per_clip = generation.get("estCostPerClip", 25)

    mode_label = f"{reference.get('mode', 'A')} — {reference.get('modeDescription', 'Single shared')}"
    rotation = reference.get("rotationStrategy", "")
    if rotation:
        mode_label += f" ({rotation})"

    summary_rows = [
        ("Project", project["name"]),
        ("Slug", project["slug"]),
        ("Vertical", project.get("vertical", "")),
        ("Model", generation["modelDisplay"]),
        ("MCP Model", generation["mcpModel"]),
        ("Duration", f"{generation['duration']}s"),
        ("Aspect ratio", generation["aspectRatio"]),
        ("Count per prompt", generation["countPerPrompt"]),
        ("Reference mode", mode_label),
        ("Total prompts", len(prompts_data)),
        ("Total clips (planned)", total_clips),
        ("Total clips delivered", delivered),
        ("Est cost", f"~{total_clips * cost_per_clip} cr ({generation['modelDisplay']} @ ~{cost_per_clip} cr/clip)"),
        ("Reference image", project.get("imgName", "")),
        ("Project fired", project.get("createdAt", "")[:10] if project.get("createdAt") else ""),
        ("Fired by", project.get("user", "")),
        ("MB", project.get("mb", "")),
        ("Notion request", project.get("notionUrl", "")),
    ]
    if project.get("notes"):
        summary_rows.append(("Notes", project["notes"]))

    for i, (label, value) in enumerate(summary_rows, start=2):
        l = summary.cell(row=i, column=1, value=label)
        l.fill = SUMMARY_LABEL_FILL; l.font = SUMMARY_LABEL_FONT
        l.alignment = Alignment(horizontal="left", vertical="center")
        l.border = BORDER
        v = summary.cell(row=i, column=2, value=value)
        v.fill = SUMMARY_VALUE_FILL
        v.alignment = Alignment(horizontal="left", vertical="center")
        v.border = BORDER

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 90

    # === Prompts sheet ===
    prompts = wb.create_sheet("Prompts")
    headers = ["L#", "Slug", "State", "Slide", "Dialogue", "Reference(s)", "Full Prompt", "Status", "v01", "v02", "Notes"]
    prompts.append(headers)
    for cell in prompts[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    prompts.row_dimensions[1].height = 26

    for idx, entry in enumerate(prompts_data):
        row_num = idx + 2
        status = entry.get("status", "pending")

        row_values = [
            entry["lineNum"],
            entry["slug"],
            entry.get("state", ""),
            entry.get("slide", ""),
            entry["dialogue"],
            entry.get("references", ""),
            entry["fullPrompt"],
            status,
            entry.get("v01", ""),
            entry.get("v02", ""),
            entry.get("notes", ""),
        ]

        for col, val in enumerate(row_values, start=1):
            c = prompts.cell(row=row_num, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if idx % 2 == 1:
                c.fill = BAND_FILL

        # Status column override (col 8)
        status_cell = prompts.cell(row=row_num, column=8)
        if status in STATUS_STYLE:
            status_fill, status_font = STATUS_STYLE[status]
            status_cell.fill = status_fill
            status_cell.font = status_font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Wrap on text-heavy short columns
        prompts.cell(row=row_num, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)   # Slide
        prompts.cell(row=row_num, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)   # Dialogue
        prompts.cell(row=row_num, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)   # Reference(s)
        prompts.cell(row=row_num, column=11).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Notes
        prompts.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")                 # L#

        # Fixed row height — keeps Full Prompt collapsed; double-click to view full JSON
        prompts.row_dimensions[row_num].height = 36

    # Column widths
    prompts.column_dimensions["A"].width = 5      # L#
    prompts.column_dimensions["B"].width = 32     # Slug
    prompts.column_dimensions["C"].width = 8      # State
    prompts.column_dimensions["D"].width = 35     # Slide
    prompts.column_dimensions["E"].width = 50     # Dialogue
    prompts.column_dimensions["F"].width = 32     # Reference(s)
    prompts.column_dimensions["G"].width = 40     # Full Prompt — narrow → text clips
    prompts.column_dimensions["H"].width = 10     # Status
    prompts.column_dimensions["I"].width = 32     # v01
    prompts.column_dimensions["J"].width = 32     # v02
    prompts.column_dimensions["K"].width = 38     # Notes

    prompts.freeze_panes = "A2"

    wb.save(output_path)

    # Stats for the runtime to report
    stats = {
        "total": len(prompts_data),
        "good": sum(1 for p in prompts_data if p["status"] == "✓"),
        "partial": sum(1 for p in prompts_data if p["status"] == "Partial"),
        "failed": sum(1 for p in prompts_data if p["status"] == "✗"),
        "pending": sum(1 for p in prompts_data if p["status"] == "pending"),
        "skipped": sum(1 for p in prompts_data if p["status"] == "skipped"),
    }
    return stats


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("Usage: build_xlsx.py <config.json> <output.xlsx>")
    config_path, output_path = sys.argv[1], sys.argv[2]
    with open(config_path) as f:
        config = json.load(f)
    stats = build(config, output_path)
    print(f"✓ Wrote {output_path}")
    print(f"  {stats['total']} rows: {stats['good']} ✓ | {stats['partial']} partial | {stats['failed']} ✗ | {stats['pending']} pending | {stats['skipped']} skipped")
