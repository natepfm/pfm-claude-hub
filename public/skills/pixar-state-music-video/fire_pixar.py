#!/usr/bin/env python3
"""
fire_pixar.py — fire a Pixar State Music Video's 13-image shot set (character + 12 scenes)
via the Higgsfield CLI (NB Pro 2k, plain prose prompts), download into the project folder,
write the Excel manifest.

Usage:
  python3 fire_pixar.py <manifest.json> <project_root> [--workers 8] [--stage /tmp/pixar_out]

manifest.json = output of build_pixar_prompts.py
project_root  = the Lucid Link project folder (…/05.18.26 - Pixar Best State Music - Oregon)

Outputs:
  Elements/Graphics/<State> PIXAR.png          (character)
  Elements/Footage/Reference/<state>_01..12.png (the 12 scene frames the editor animates)
  Elements/Prompts/pixar_<state>_<vertical>_shots.xlsx   (manifest)

NB Pro takes PLAIN PROSE prompts — a JSON-blob prompt returns HTTP 500 (learned 2026-05-30).
Retry caps mirror the house rule: 1 + up to 2 retries on transient 500/timeout, then flag.
"""
import subprocess, json, os, sys, shutil, urllib.request, argparse
from concurrent.futures import ThreadPoolExecutor, as_completed

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("manifest"); ap.add_argument("project_root")
    ap.add_argument("--workers",type=int,default=8); ap.add_argument("--stage",default="/tmp/pixar_out")
    ap.add_argument("--log",default="/tmp/pixar_fire.log")
    a=ap.parse_args()
    M=json.load(open(a.manifest)); P=a.project_root
    GFX=os.path.join(P,"Elements/Graphics"); REF=os.path.join(P,"Elements/Footage/Reference")
    PR=os.path.join(P,"Elements/Prompts")
    for d in (GFX,REF,PR,a.stage): os.makedirs(d,exist_ok=True)
    open(a.log,"w").close()
    def log(m): open(a.log,"a").write(m+"\n"); print(m,flush=True)

    def parse(out):
        try: o=json.loads(out)
        except:
            ls=[l for l in out.splitlines() if l.strip().startswith(("[","{"))]; o=json.loads(ls[-1]) if ls else {}
        if isinstance(o,list): o=o[0] if o else {}
        return (o.get("result_url") or o.get("url")),(o.get("status") if isinstance(o,dict) else None)

    def fire(shot):
        out_dir = GFX if shot["kind"]=="character" else REF
        dest=os.path.join(out_dir,shot["outName"]); stg=os.path.join(a.stage,shot["outName"])
        for att in range(1,4):  # initial + 2 retries (transient 500s)
            try:
                r=subprocess.run(["higgsfield","generate","create","nano_banana_2","--prompt",shot["prompt"],
                    "--aspect_ratio","9:16","--resolution","2k","--wait","--wait-timeout","4m","--json"],
                    capture_output=True,text=True,timeout=280)
                if r.returncode!=0: log(f"  {shot['shotId']} a{att}: {r.stderr.strip()[:70]}"); continue
                u,st=parse(r.stdout)
                if not u: log(f"  {shot['shotId']} a{att}: status={st}"); continue
                req=urllib.request.Request(u,headers={"User-Agent":"Mozilla/5.0"})
                with urllib.request.urlopen(req,timeout=120) as resp,open(stg,"wb") as fh: fh.write(resp.read())
                if os.path.getsize(stg)>2000:
                    shutil.copy2(stg,dest); return (shot["shotId"],"ok")
            except Exception as e: log(f"  {shot['shotId']} a{att}: {type(e).__name__}")
        return (shot["shotId"],"FAILED")

    shots=M["shots"]
    log(f"FIRING {len(shots)} Pixar shots ({M['project']['name']}) — NB Pro 2k prose")
    results={}
    with ThreadPoolExecutor(max_workers=a.workers) as ex:
        for f in as_completed({ex.submit(fire,s):s for s in shots}):
            sid,res=f.result(); results[sid]=res; log(f"{'OK' if res=='ok' else 'FAIL'} {sid}")
    ok=sum(1 for v in results.values() if v=="ok"); bad=[k for k,v in results.items() if v!="ok"]

    # --- xlsx manifest ---
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Shots"
        ws.append(["Shot ID","Kind","Scene #","Out File","Status","Prompt (prose)"])
        for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="305496")
        for s in shots:
            ws.append([s["shotId"],s["kind"],s.get("sceneNum",""),s["outName"],
                       "✓" if results.get(s["shotId"])=="ok" else "✗",s["prompt"][:300]])
        for col,w in zip("ABCDEF",[26,10,8,34,8,80]): ws.column_dimensions[col].width=w
        sm=wb.create_sheet("Summary")
        for k,v in [("Project",M["project"]["name"]),("State",M["project"]["state"]),
                    ("Vertical",M["project"]["vertical"]),("Reveal",M["project"]["revealRate"]),
                    ("Shots",f"{ok}/{len(shots)} delivered"),("Model","Nano Banana Pro 2k, 9:16, count=1"),
                    ("Note","Character -> Graphics/, 12 scenes -> Footage/Reference/. Editor animates frames + does Suno + Mark edits.")]:
            sm.append([k,v])
        xlsx=os.path.join(PR,f"pixar_{M['project']['slug']}_shots.xlsx")
        wb.save(xlsx); log(f"manifest: {xlsx}")
    except Exception as e: log(f"xlsx skip: {e}")

    log(f"\nDONE: {ok}/{len(shots)} delivered | FAILED: {bad if bad else 'none'}")
    json.dump({"ok":ok,"failed":bad},open("/tmp/pixar_fire_result.json","w"))

if __name__=="__main__": main()
