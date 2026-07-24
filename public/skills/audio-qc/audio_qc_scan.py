#!/usr/bin/env python3
"""Audio QC scanner for Veo-generated mp4 clips.

Two phases:
  Phase 1 — ffmpeg audio-physics checks (parallel, ~90s per ~350 clips)
    1. Audio stream present?
    2. mean_volume > -55 dB ?     (silent stream)
    3. max_volume < -0.5 dB ?     (clipping)

  Phase 2 (optional, requires --manifest) — Whisper dialogue verification
    (sequential, ~0.3s per clip on M-series Mac, so ~2 min for 350 clips)
    - Loads Whisper model once
    - Transcribes each clip
    - Fuzzy-matches transcript against expected dialogue from the project's
      Excel manifest. Flags `dialogue_mismatch` on clips below the threshold
      (default 0.70 similarity) — likely wrong words / voice drift.

Cut-off heuristic was DISABLED 2026-05-26 (false-positive on Rachel's
continuous-narration style — see "What's NOT in this skill" in the audio-qc
SKILL.md). Tail-silence is still computed and reported for visibility but
no longer sets a flag.

Whisper folded back in 2026-05-26 after the M-series speed test showed
0.3s/clip (vs the 10-15s/clip the original handoff doc estimated) — the
full pass is now fast enough to run by default whenever a manifest is
available.

Flag values: OK | silent | low_volume | clipped | no_audio | dialogue_mismatch | error

Usage:
    audio_qc_scan.py <veo_root> [--out <report.md>] [--exclude SUBSTR]
                     [--workers N] [--manifest <xlsx>]
                     [--whisper-model base|small|...] [--no-whisper]
                     [--dialogue-threshold 0.70]
"""

from __future__ import annotations  # PEP 563 — for `str | None` style on Python 3.9

import argparse
import difflib
import os
import re
import shutil
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path


def _resolve_ffmpeg() -> str:
    """Resolve ffmpeg binary path. Order: $PFM_FFMPEG env → ~/bin/ffmpeg → PATH."""
    env = os.environ.get("PFM_FFMPEG")
    if env and os.path.isfile(env):
        return env
    home_bin = os.path.expanduser("~/bin/ffmpeg")
    if os.path.isfile(home_bin):
        return home_bin
    on_path = shutil.which("ffmpeg")
    if on_path:
        return on_path
    print(
        "error: ffmpeg not found. Install via the PFM setup script:\n"
        '  bash "/Volumes/ads/PFM MEDIA MASTER FOLDER/6. Claude PFM/claude-pfm-setup.sh"\n'
        "or set PFM_FFMPEG to its absolute path.",
        file=sys.stderr,
    )
    sys.exit(2)


FFMPEG = _resolve_ffmpeg()

# Thresholds — audio-physics
MEAN_VOL_SILENT_DB = -55.0     # below this => silent
MEAN_VOL_LOW_DB = -35.0        # below this (but above silent) => low_volume
MAX_VOL_CLIP_DB = -0.5         # above this => clipped
TAIL_ZONE_S = 0.5              # last N seconds of clip checked for silence (reported, not flagged)
SILENCEDETECT_DB = -35
SILENCEDETECT_MIN_S = 0.15

# Thresholds — Whisper dialogue match
DIALOGUE_MATCH_DEFAULT_THRESHOLD = 0.70
WHISPER_MODEL_DEFAULT = "base"

# Thresholds — Phase 3 unexpected music / non-speech audio detection
# (runs whenever Phase 2 / Whisper runs, since it needs speech segments)
MUSIC_MIN_REGION_S = 0.30          # drop regions shorter than this (noise / breath / Whisper word-gap artifacts)
MUSIC_FLAG_SINGLE_REGION_S = 0.40  # any single non-speech region above this → flag
MUSIC_FLAG_TOTAL_S = 0.60          # cumulative non-speech audio above this → flag

DURATION_RE = re.compile(r"Duration:\s*(\d+):(\d+):(\d+\.\d+)")
AUDIO_STREAM_RE = re.compile(r"Stream #0:\d+.*Audio:")
MEAN_VOL_RE = re.compile(r"mean_volume:\s*(-?\d+\.?\d*)\s*dB")
MAX_VOL_RE = re.compile(r"max_volume:\s*(-?\d+\.?\d*)\s*dB")
SILENCE_START_RE = re.compile(r"silence_start:\s*(-?\d+\.?\d*)")
SILENCE_END_RE = re.compile(r"silence_end:\s*(-?\d+\.?\d*)")

PLACEHOLDER_TOKEN_RE = re.compile(r"\[[A-Z][A-Z0-9_]*\]")  # [STATE], [STATE_ANNUAL_RATE], etc.
SLUG_FROM_FILENAME_RE = re.compile(r"^(.+?)_v\d+\.mp4$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Phase 1 — ffmpeg audio-physics checks
# ---------------------------------------------------------------------------

def probe_clip(file: Path) -> dict:
    """Run ffmpeg checks on one clip; return parsed result dict."""
    out = {
        "file": file,
        "duration_s": None,
        "audio_stream": False,
        "mean_vol_dB": None,
        "max_vol_dB": None,
        "tail_silence_s": None,
        "silence_intervals": [],  # list of (start, end) tuples — feeds Phase 3
        "flag": "OK",
        "error": None,
        # Phase 2 fields — filled later if Whisper runs
        "transcript": None,
        "expected_dialogue": None,
        "dialogue_similarity": None,
        # Phase 3 fields — filled later if Whisper ran (needs speech segments)
        "speech_segments": None,        # list of (start, end) tuples from Whisper
        "music_regions": None,          # list of (start, end) — non-silent non-speech
        "music_total_s": None,          # cumulative seconds of music_regions
        "has_unexpected_music": False,  # True if above MUSIC_FLAG_* thresholds
    }

    try:
        result = subprocess.run(
            [
                FFMPEG, "-hide_banner", "-nostats", "-i", str(file),
                "-af", (f"volumedetect,silencedetect="
                        f"noise={SILENCEDETECT_DB}dB:"
                        f"duration={SILENCEDETECT_MIN_S}"),
                "-vn", "-sn", "-dn", "-f", "null", "-",
            ],
            capture_output=True, text=True, timeout=60,
        )
        stderr = result.stderr
    except subprocess.TimeoutExpired:
        out["flag"] = "error"
        out["error"] = "ffmpeg timeout"
        return out
    except Exception as e:
        out["flag"] = "error"
        out["error"] = f"ffmpeg exception: {e}"
        return out

    m = DURATION_RE.search(stderr)
    if m:
        h, mn, s = m.groups()
        out["duration_s"] = int(h) * 3600 + int(mn) * 60 + float(s)

    out["audio_stream"] = bool(AUDIO_STREAM_RE.search(stderr))

    m = MEAN_VOL_RE.search(stderr)
    if m:
        out["mean_vol_dB"] = float(m.group(1))
    m = MAX_VOL_RE.search(stderr)
    if m:
        out["max_vol_dB"] = float(m.group(1))

    silence_starts = [float(x) for x in SILENCE_START_RE.findall(stderr)]
    silence_ends = [float(x) for x in SILENCE_END_RE.findall(stderr)]

    # Pair up silence intervals (pad ends with duration if shorter — silence at end of clip)
    silence_intervals: list[tuple[float, float]] = []
    if out["duration_s"] is not None:
        for i, s in enumerate(silence_starts):
            e = silence_ends[i] if i < len(silence_ends) else out["duration_s"]
            silence_intervals.append((s, e))
    out["silence_intervals"] = silence_intervals

    tail_silence_s = 0.0
    if out["duration_s"] is not None:
        tail_zone_start = out["duration_s"] - TAIL_ZONE_S
        for s, e in silence_intervals:
            overlap = max(0.0, min(e, out["duration_s"]) - max(s, tail_zone_start))
            if overlap > tail_silence_s:
                tail_silence_s = overlap
    out["tail_silence_s"] = tail_silence_s

    # Decide audio-physics flag, priority order
    if not out["audio_stream"]:
        out["flag"] = "no_audio"
    elif out["mean_vol_dB"] is None:
        out["flag"] = "error"
        out["error"] = "could not parse mean_volume"
    elif out["mean_vol_dB"] < MEAN_VOL_SILENT_DB:
        out["flag"] = "silent"
    elif out["mean_vol_dB"] < MEAN_VOL_LOW_DB:
        out["flag"] = "low_volume"
    elif out["max_vol_dB"] is not None and out["max_vol_dB"] > MAX_VOL_CLIP_DB:
        out["flag"] = "clipped"
    # cut_off heuristic DISABLED 2026-05-26 — see header docstring.

    return out


# ---------------------------------------------------------------------------
# Phase 2 — Whisper dialogue verification
# ---------------------------------------------------------------------------

def load_manifest_dialogue(manifest_path: Path) -> dict[str, str]:
    """Read the Excel manifest, return {slug: expected_dialogue} dict.

    Looks for the "Prompts" sheet (per hvg-flow's build_xlsx.py convention).
    Identifies `slug` and `dialogue` columns by case-insensitive header match.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "error: openpyxl not installed. Install via the PFM setup script:\n"
            '  bash "/Volumes/ads/PFM MEDIA MASTER FOLDER/6. Claude PFM/claude-pfm-setup.sh"\n'
            "or: pip3 install --user openpyxl",
            file=sys.stderr,
        )
        sys.exit(2)

    wb = load_workbook(manifest_path, read_only=True, data_only=True)
    sheet_name = "Prompts" if "Prompts" in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[sheet_name]

    rows = sheet.iter_rows(values_only=True)
    headers = [str(h or "").strip().lower() for h in next(rows, [])]

    try:
        slug_idx = headers.index("slug")
    except ValueError:
        print(f"error: manifest missing 'slug' column. Headers: {headers}", file=sys.stderr)
        sys.exit(2)
    try:
        dialogue_idx = headers.index("dialogue")
    except ValueError:
        print(f"error: manifest missing 'dialogue' column. Headers: {headers}", file=sys.stderr)
        sys.exit(2)

    mapping = {}
    for row in rows:
        if not row or len(row) <= max(slug_idx, dialogue_idx):
            continue
        slug = row[slug_idx]
        dialogue = row[dialogue_idx]
        if slug and dialogue:
            mapping[str(slug).strip()] = str(dialogue).strip()

    return mapping


def extract_slug(file: Path) -> str | None:
    """Strip `_v\\d+.mp4` from filename to recover the manifest slug."""
    m = SLUG_FROM_FILENAME_RE.match(file.name)
    if m:
        return m.group(1)
    return None


def normalize_for_match(text: str) -> str:
    """Normalize text for fuzzy matching: drop placeholders, lowercase,
    strip punctuation, collapse whitespace."""
    text = PLACEHOLDER_TOKEN_RE.sub("", text)
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def similarity(expected: str, transcript: str) -> float:
    """Return SequenceMatcher.ratio() on normalized strings (0.0-1.0)."""
    a = normalize_for_match(expected)
    b = normalize_for_match(transcript)
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def _invert_intervals(
    intervals: list[tuple[float, float]],
    range_start: float,
    range_end: float,
) -> list[tuple[float, float]]:
    """Return the complement of `intervals` within [range_start, range_end]."""
    if range_end <= range_start:
        return []
    if not intervals:
        return [(range_start, range_end)]

    sorted_iv = sorted(intervals, key=lambda x: x[0])
    result: list[tuple[float, float]] = []
    cursor = range_start
    for s, e in sorted_iv:
        s = max(s, range_start)
        e = min(e, range_end)
        if s > cursor:
            result.append((cursor, s))
        cursor = max(cursor, e)
    if cursor < range_end:
        result.append((cursor, range_end))
    return result


def _subtract_interval(
    regions: list[tuple[float, float]],
    sub_start: float,
    sub_end: float,
) -> list[tuple[float, float]]:
    """Remove [sub_start, sub_end] from each region in `regions`."""
    result: list[tuple[float, float]] = []
    for r_start, r_end in regions:
        # No overlap
        if sub_end <= r_start or sub_start >= r_end:
            result.append((r_start, r_end))
            continue
        # Overlap — may leave 0, 1, or 2 sub-regions
        if sub_start > r_start:
            result.append((r_start, sub_start))
        if sub_end < r_end:
            result.append((sub_end, r_end))
    return result


def detect_unexpected_music(
    silence_intervals: list[tuple[float, float]],
    speech_segments: list[tuple[float, float]],
    duration: float,
    min_region_s: float = MUSIC_MIN_REGION_S,
) -> list[tuple[float, float]]:
    """Return list of (start, end) regions where there's audio energy but
    Whisper detected no speech — i.e. non-silent non-speech audio.

    Logic:
      1. Compute non-silent regions (invert silence_intervals across [0, duration])
      2. Subtract speech segments — what's left is non-speech non-silent audio
      3. Filter regions shorter than min_region_s (drops breath / Whisper word-gap noise)

    In a dialogue-expected clip, this surfaces musical stings, musical beds,
    musical tails, or any non-speech audio that shouldn't be there.

    Caveat: if Whisper missed any spoken words (silent voice, muffled audio,
    unusual accent), they'd appear here as false-positive music. The
    min_region_s filter mitigates short bursts. Spot-check flagged clips
    before concluding "music."
    """
    if duration is None or duration <= 0:
        return []

    non_silent = _invert_intervals(silence_intervals, 0.0, duration)
    candidates = non_silent
    for seg_start, seg_end in speech_segments:
        candidates = _subtract_interval(candidates, seg_start, seg_end)

    return [(s, e) for s, e in candidates if (e - s) >= min_region_s]


def run_whisper_phase(
    results: list[dict],
    manifest: dict[str, str],
    model_name: str,
    threshold: float,
) -> None:
    """Mutate `results` in place: add transcript / expected_dialogue /
    dialogue_similarity per clip. If audio-physics flag is OK AND similarity
    < threshold, escalate flag to 'dialogue_mismatch'.

    Skips clips with audio-physics flag of `silent`, `no_audio`, or `error`
    (no point transcribing those).
    """
    try:
        import whisper  # type: ignore
    except ImportError:
        print(
            "error: openai-whisper not installed. Install via the PFM setup script:\n"
            '  bash "/Volumes/ads/PFM MEDIA MASTER FOLDER/6. Claude PFM/claude-pfm-setup.sh"\n'
            "or: pip3 install --user -U openai-whisper",
            file=sys.stderr,
        )
        sys.exit(2)

    print(f"\nLoading Whisper model '{model_name}'...", file=sys.stderr)
    t0 = time.time()
    model = whisper.load_model(model_name)
    print(f"  model loaded in {time.time() - t0:.1f}s", file=sys.stderr)

    transcribable = [r for r in results if r["flag"] not in ("silent", "no_audio", "error")]
    print(f"Transcribing {len(transcribable)} clips (skipping "
          f"{len(results) - len(transcribable)} with audio-physics failures)...",
          file=sys.stderr)

    for i, r in enumerate(transcribable, 1):
        slug = extract_slug(r["file"])
        if not slug or slug not in manifest:
            r["expected_dialogue"] = None
            r["transcript"] = "(no manifest entry — skipped)"
            r["dialogue_similarity"] = None
            continue

        r["expected_dialogue"] = manifest[slug]
        try:
            result = model.transcribe(str(r["file"]), fp16=False, language="en", verbose=False)
            transcript = result.get("text", "").strip()
            r["transcript"] = transcript
            r["dialogue_similarity"] = similarity(r["expected_dialogue"], transcript)

            if r["flag"] == "OK" and r["dialogue_similarity"] < threshold:
                r["flag"] = "dialogue_mismatch"

            # Phase 3 — unexpected music / non-speech audio detection.
            # Piggybacks on Phase 1's silence_intervals + Whisper's speech segments.
            r["speech_segments"] = [
                (float(seg["start"]), float(seg["end"]))
                for seg in result.get("segments", [])
            ]
            if r["duration_s"] and r["flag"] not in ("silent", "no_audio", "error"):
                music_regions = detect_unexpected_music(
                    silence_intervals=r["silence_intervals"],
                    speech_segments=r["speech_segments"],
                    duration=r["duration_s"],
                )
                r["music_regions"] = music_regions
                r["music_total_s"] = sum(e - s for s, e in music_regions)
                has_significant_music = (
                    any((e - s) > MUSIC_FLAG_SINGLE_REGION_S for s, e in music_regions)
                    or r["music_total_s"] > MUSIC_FLAG_TOTAL_S
                )
                r["has_unexpected_music"] = has_significant_music
                # Escalate flag only if currently OK — don't override
                # dialogue_mismatch / clipped / low_volume (those are still
                # primary issues; music shows up in its own report section too).
                if has_significant_music and r["flag"] == "OK":
                    r["flag"] = "unexpected_music"
        except Exception as e:
            r["transcript"] = f"(whisper error: {e})"
            r["dialogue_similarity"] = None

        if i % 25 == 0 or i == len(transcribable):
            print(f"  {i}/{len(transcribable)}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Report rendering
# ---------------------------------------------------------------------------

def parse_path(file: Path, veo_root: Path) -> tuple[str, str, str]:
    """Extract (batch, state, clip_name) from path relative to veo_root."""
    try:
        rel = file.relative_to(veo_root)
        parts = rel.parts
        if len(parts) == 1:
            return ("(root)", "", parts[0])
        if len(parts) == 2:
            return (parts[0], "", parts[1])
        return (parts[0], parts[1], parts[-1])
    except ValueError:
        return ("?", "?", file.name)


def fmt(v, spec="{:.2f}"):
    return spec.format(v) if v is not None else "—"


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("veo_root", help="Root of Elements/Footage/Veo/")
    p.add_argument("--out", default=None, help="Output markdown path")
    p.add_argument("--exclude", action="append", default=[],
                   help="Skip files whose path contains this substring (repeatable)")
    p.add_argument("--workers", type=int, default=12,
                   help="Phase 1 ffmpeg worker count (default 12)")
    p.add_argument("--manifest", default=None,
                   help="Path to project's Excel manifest (enables Whisper phase)")
    p.add_argument("--whisper-model", default=WHISPER_MODEL_DEFAULT,
                   help=f"Whisper model name (default {WHISPER_MODEL_DEFAULT})")
    p.add_argument("--no-whisper", action="store_true",
                   help="Skip Whisper phase even if --manifest is provided")
    p.add_argument("--dialogue-threshold", type=float, default=DIALOGUE_MATCH_DEFAULT_THRESHOLD,
                   help=f"Min similarity to clear (default {DIALOGUE_MATCH_DEFAULT_THRESHOLD})")
    args = p.parse_args()

    veo_root = Path(args.veo_root).resolve()
    if not veo_root.is_dir():
        print(f"error: {veo_root} not a directory", file=sys.stderr)
        sys.exit(1)

    all_files = sorted(veo_root.rglob("*.mp4"))
    files = [f for f in all_files if not any(ex in str(f) for ex in args.exclude)]
    skipped = len(all_files) - len(files)

    print(f"Found {len(all_files)} mp4 files; scanning {len(files)} "
          f"(skipped {skipped})", file=sys.stderr)
    if not files:
        sys.exit(0)

    # ---- Phase 1: ffmpeg parallel ----
    print(f"\nPhase 1 — ffmpeg audio-physics checks (workers={args.workers})...",
          file=sys.stderr)
    t_phase1 = time.time()
    results: list[dict] = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(probe_clip, f): f for f in files}
        for i, fut in enumerate(as_completed(futs), 1):
            results.append(fut.result())
            if i % 25 == 0 or i == len(files):
                print(f"  {i}/{len(files)}", file=sys.stderr)
    print(f"Phase 1 done in {time.time() - t_phase1:.1f}s", file=sys.stderr)

    results.sort(key=lambda r: str(r["file"]))

    # ---- Phase 2: Whisper (optional) ----
    whisper_ran = False
    manifest_mapping = {}
    if args.manifest and not args.no_whisper:
        manifest_path = Path(args.manifest).expanduser().resolve()
        if not manifest_path.is_file():
            # G1 gate: a requested-but-unreadable manifest must NOT silently degrade to physics-only —
            # that reports "QC clean" on a scan that never checked the dialogue.
            print(f"ERROR: --manifest given but not found at {manifest_path}. "
                  f"Fix the path (or run explicitly without --manifest for a physics-only scan). Exiting.",
                  file=sys.stderr)
            sys.exit(2)
        else:
            print(f"\nPhase 2 — Whisper dialogue verification (manifest: {manifest_path.name})...",
                  file=sys.stderr)
            t_phase2 = time.time()
            manifest_mapping = load_manifest_dialogue(manifest_path)
            print(f"  manifest entries: {len(manifest_mapping)}", file=sys.stderr)
            run_whisper_phase(results, manifest_mapping, args.whisper_model,
                              args.dialogue_threshold)
            print(f"Phase 2 done in {time.time() - t_phase2:.1f}s", file=sys.stderr)
            whisper_ran = True

    # ---- Report ----
    flag_counts: dict[str, int] = {}
    for r in results:
        flag_counts[r["flag"]] = flag_counts.get(r["flag"], 0) + 1

    out_path = (Path(args.out) if args.out
                else veo_root / f"audio_qc_report_{datetime.now():%Y-%m-%d}.md")

    lines: list[str] = []
    lines.append(f"# Audio QC — {datetime.now():%Y-%m-%d %H:%M}")
    lines.append("")
    lines.append(f"- Root: `{veo_root}`")
    lines.append(f"- Scanned: **{len(results)}** mp4 files")
    if args.exclude:
        lines.append(f"- Excluded patterns: {', '.join(args.exclude)} (skipped {skipped})")
    lines.append(f"- Audio thresholds: silent mean < {MEAN_VOL_SILENT_DB} dB · "
                 f"low_volume mean < {MEAN_VOL_LOW_DB} dB · "
                 f"clipped max > {MAX_VOL_CLIP_DB} dB")
    if whisper_ran:
        lines.append(f"- Whisper model: `{args.whisper_model}` · "
                     f"dialogue match threshold: {args.dialogue_threshold}")
    lines.append("")

    lines.append("## Summary")
    lines.append("")
    lines.append("| flag | count |")
    lines.append("|---|---|")
    for flag in ("OK", "silent", "low_volume", "clipped", "no_audio",
                 "dialogue_mismatch", "unexpected_music", "error"):
        if flag in flag_counts:
            lines.append(f"| {flag} | {flag_counts[flag]} |")
    lines.append("")

    flagged = [r for r in results if r["flag"] != "OK"]
    lines.append(f"## Flagged clips ({len(flagged)})")
    lines.append("")
    if flagged:
        if whisper_ran:
            lines.append("| batch | state | clip | dur | mean dB | max dB | "
                         "sim | flag |")
            lines.append("|---|---|---|---|---|---|---|---|")
            for r in flagged:
                batch, state, clip = parse_path(r["file"], veo_root)
                lines.append(
                    f"| {batch} | {state} | `{clip}` | "
                    f"{fmt(r['duration_s'])} | "
                    f"{fmt(r['mean_vol_dB'], '{:.1f}')} | "
                    f"{fmt(r['max_vol_dB'], '{:.1f}')} | "
                    f"{fmt(r['dialogue_similarity'])} | "
                    f"**{r['flag']}** |"
                )
        else:
            lines.append("| batch | state | clip | dur | mean dB | max dB | flag |")
            lines.append("|---|---|---|---|---|---|---|")
            for r in flagged:
                batch, state, clip = parse_path(r["file"], veo_root)
                lines.append(
                    f"| {batch} | {state} | `{clip}` | "
                    f"{fmt(r['duration_s'])} | "
                    f"{fmt(r['mean_vol_dB'], '{:.1f}')} | "
                    f"{fmt(r['max_vol_dB'], '{:.1f}')} | "
                    f"**{r['flag']}** |"
                )
    else:
        lines.append("_None — all clips passed._")
    lines.append("")

    # ---- Dialogue mismatch detail (if Whisper ran) ----
    if whisper_ran:
        mismatches = [r for r in results if r["flag"] == "dialogue_mismatch"]
        lines.append(f"## Dialogue mismatches — transcript detail ({len(mismatches)})")
        lines.append("")
        if mismatches:
            for r in mismatches:
                batch, state, clip = parse_path(r["file"], veo_root)
                lines.append(f"### `{clip}` ({batch}/{state}, sim {fmt(r['dialogue_similarity'])})")
                lines.append("")
                lines.append(f"**Expected:** {r['expected_dialogue']}")
                lines.append("")
                lines.append(f"**Heard:** {r['transcript']}")
                lines.append("")
        else:
            lines.append("_None — every transcribed clip matched expected dialogue above threshold._")
            lines.append("")

    # ---- Unexpected music detail (Phase 3, if Whisper ran) ----
    if whisper_ran:
        music_clips = [r for r in results if r.get("has_unexpected_music")]
        lines.append(f"## Unexpected music / non-speech audio ({len(music_clips)})")
        lines.append("")
        if music_clips:
            lines.append("Non-speech audio detected — likely musical sting (cold-open or tail), musical bed, ")
            lines.append("or other non-dialogue energy. Whisper transcribed the speech, but audio energy ")
            lines.append("remains in regions where no words were spoken.")
            lines.append("")
            lines.append(f"_Thresholds: flag if any single region > {MUSIC_FLAG_SINGLE_REGION_S}s OR cumulative > {MUSIC_FLAG_TOTAL_S}s. "
                         f"Regions shorter than {MUSIC_MIN_REGION_S}s filtered out as noise / breath._")
            lines.append("")
            lines.append("| batch | state | clip | music total s | regions (s) | primary flag |")
            lines.append("|---|---|---|---|---|---|")
            for r in music_clips:
                batch, state, clip = parse_path(r["file"], veo_root)
                regions_str = ", ".join(
                    f"{s:.2f}–{e:.2f}" for s, e in (r["music_regions"] or [])
                )
                lines.append(
                    f"| {batch} | {state} | `{clip}` | "
                    f"{fmt(r['music_total_s'])} | "
                    f"{regions_str} | "
                    f"**{r['flag']}** |"
                )
            lines.append("")
            lines.append("_Caveat: if Whisper missed any spoken words (silent voice, muffled audio, unusual accent), "
                         "they'll appear here as false-positive music. Spot-check before refiring._")
            lines.append("")
        else:
            lines.append("_None — no significant non-speech audio detected._")
            lines.append("")

    # ---- Full table ----
    lines.append("## All clips")
    lines.append("")
    lines.append("<details><summary>Expand full results</summary>")
    lines.append("")
    if whisper_ran:
        lines.append("| batch | state | clip | dur | mean dB | max dB | "
                     "tail silence s | sim | flag |")
        lines.append("|---|---|---|---|---|---|---|---|---|")
        for r in results:
            batch, state, clip = parse_path(r["file"], veo_root)
            lines.append(
                f"| {batch} | {state} | `{clip}` | "
                f"{fmt(r['duration_s'])} | "
                f"{fmt(r['mean_vol_dB'], '{:.1f}')} | "
                f"{fmt(r['max_vol_dB'], '{:.1f}')} | "
                f"{fmt(r['tail_silence_s'])} | "
                f"{fmt(r['dialogue_similarity'])} | "
                f"{r['flag']} |"
            )
    else:
        lines.append("| batch | state | clip | dur | mean dB | max dB | "
                     "tail silence s | flag |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for r in results:
            batch, state, clip = parse_path(r["file"], veo_root)
            lines.append(
                f"| {batch} | {state} | `{clip}` | "
                f"{fmt(r['duration_s'])} | "
                f"{fmt(r['mean_vol_dB'], '{:.1f}')} | "
                f"{fmt(r['max_vol_dB'], '{:.1f}')} | "
                f"{fmt(r['tail_silence_s'])} | "
                f"{r['flag']} |"
            )
    lines.append("")
    lines.append("</details>")
    lines.append("")

    out_path.write_text("\n".join(lines) + "\n")
    print(f"\nReport: {out_path}", file=sys.stderr)
    print(f"Flag counts: {flag_counts}", file=sys.stderr)


if __name__ == "__main__":
    main()
