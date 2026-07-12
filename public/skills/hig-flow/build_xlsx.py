#!/usr/bin/env python3
"""
HIG flow — build the project audit spreadsheet for image generations.

Usage:
    python3 build_xlsx.py <config.json> <output.xlsx>

The config JSON shape:
{
  "project": {
    "name":      "<canonical Notion project name>",
    "slug":      "<derived slug used in filenames>",
    "vertical":  "Auto - Forms",
    "createdAt": "2026-05-07",
    "user":      "Sam",
    "mb":        "David V",
    "notionUrl": "https://www.notion.so/...",
    "notes":     "optional context line shown on Summary sheet"
  },
  "generation": {
    "modelDisplay":     "Nano Banana Pro",
    "mcpModel":         "nano_banana_2",
    "resolution":       "1k",
    "countPerPrompt":   1,
    "estCostPerImage":  5
  },
  "shots": [
    {
      "shotId":      "L17_ryan_portrait",
      "lineMatch":   17,
      "description": "<one-line plain English>",
      "references":  ["Reference/Ryan - Master.png"],
      "aspectRatio": "9:16",
      "style":       "camera-roll",
      "prompt":      "<full Higgsfield prompt as a string>",
      "status":      "pending" | "✓" | "Partial" | "✗" | "skipped",
      "v01":         "<filename or empty>",
      "v02":         "<filename or empty>",
      "notes":       ""
    }
  ]
}

Status values render with color:
  ✓        → green
  Partial  → yellow
  ✗        → red
  pending  → grey
  skipped  → grey (italic)
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# === Style palette (matches hvg-flow for visual consistency) ===
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_VALUE_FILL = PatternFill("solid", fgColor="DDEBF7")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE"); GOOD_FONT = Font(color="006100", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFEB9C"); WARN_FONT = Font(color="9C5700", bold=True)
BAD_FILL  = PatternFill("solid", fgColor="FFC7CE"); BAD_FONT  = Font(color="9C0006", bold=True)
PEND_FILL = PatternFill("solid", fgColor="E7E6E6"); PEND_FONT = Font(color="595959", bold=True)
SKIP_FILL = PatternFill("solid", fgColor="E7E6E6"); SKIP_FONT = Font(color="595959", italic=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_STYLE = {
    "✓":       (GOOD_FILL, GOOD_FONT),
    "Partial": (WARN_FILL, WARN_FONT),
    "✗":       (BAD_FILL,  BAD_FONT),
    "pending": (PEND_FILL, PEND_FONT),
    "skipped": (SKIP_FILL, SKIP_FONT),
}


def build(config: dict, output_path: str) -> dict:
    project = config["project"]
    generation = config["generation"]
    shots_data = config["shots"]

    wb = Workbook()

    # === Summary sheet ===
    summary = wb.active
    summary.title = "Summary"

    summary.merge_cells("A1:B1")
    title_cell = summary["A1"]
    title_cell.value = "HIG Flow — Image Project Audit"
    title_cell.fill = HEADER_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 28

    total_images = len(shots_data) * generation["countPerPrompt"]
    delivered = sum(1 for s in shots_data if s["status"] == "✓") * generation["countPerPrompt"]
    delivered += sum(1 for s in shots_data if s["status"] == "Partial")  # 1 of 2

    cost_per_image = generation.get("estCostPerImage", 5)

    summary_rows = [
        ("Project", project["name"]),
        ("Slug", project["slug"]),
        ("Vertical", project.get("vertical", "")),
        ("Model", generation["modelDisplay"]),
        ("MCP Model", generation["mcpModel"]),
        ("Resolution", generation.get("resolution", "1k")),
        ("Count per prompt", generation["countPerPrompt"]),
        ("Total shots", len(shots_data)),
        ("Total images (planned)", total_images),
        ("Total images delivered", delivered),
        ("Est cost", f"~{total_images * cost_per_image} cr ({generation['modelDisplay']} @ ~{cost_per_image} cr/image)"),
        ("Created", project.get("createdAt", "")[:10] if project.get("createdAt") else ""),
        ("Fired by", project.get("user", "")),
        ("MB", project.get("mb", "")),
        ("Notion request", project.get("notionUrl", "")),
    ]
    if project.get("notes"):
        summary_rows.append(("Notes", project["notes"]))

    for i, (label, value) in enumerate(summary_rows, start=2):
        l = summary.cell(row=i, column=1, value=label)
        l.fill = SUMMARY_LABEL_FILL; l.font = SUMMARY_LABEL_FONT
        l.alignment = Alignment(horizontal="left", vertical="center")
        l.border = BORDER
        v = summary.cell(row=i, column=2, value=value)
        v.fill = SUMMARY_VALUE_FILL
        v.alignment = Alignment(horizontal="left", vertical="center")
        v.border = BORDER

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 90

    # === Shots sheet ===
    shots = wb.create_sheet("Shots")
    headers = ["Shot ID", "L#", "Description", "References", "Aspect", "Style", "Prompt", "Status", "v01", "v02", "Notes"]
    shots.append(headers)
    for cell in shots[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    shots.row_dimensions[1].height = 26

    for idx, entry in enumerate(shots_data):
        row_num = idx + 2
        status = entry.get("status", "pending")

        refs = entry.get("references", [])
        if isinstance(refs, list):
            refs_display = "\n".join(refs)
        else:
            refs_display = str(refs)

        line_match = entry.get("lineMatch", "")

        row_values = [
            entry["shotId"],
            line_match if line_match else "",
            entry.get("description", ""),
            refs_display,
            entry.get("aspectRatio", ""),
            entry.get("style", ""),
            entry["prompt"],
            status,
            entry.get("v01", ""),
            entry.get("v02", ""),
            entry.get("notes", ""),
        ]

        for col, val in enumerate(row_values, start=1):
            c = shots.cell(row=row_num, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if idx % 2 == 1:
                c.fill = BAND_FILL

        # Status column override (col 8)
        status_cell = shots.cell(row=row_num, column=8)
        if status in STATUS_STYLE:
            status_fill, status_font = STATUS_STYLE[status]
            status_cell.fill = status_fill
            status_cell.font = status_font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Wrap text on text-heavy short columns
        shots.cell(row=row_num, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)   # Description
        shots.cell(row=row_num, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)   # References (multi-line)
        shots.cell(row=row_num, column=11).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Notes
        shots.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")                 # L#
        shots.cell(row=row_num, column=5).alignment = Alignment(horizontal="center", vertical="center")                 # Aspect
        shots.cell(row=row_num, column=6).alignment = Alignment(horizontal="center", vertical="center")                 # Style

        # Fixed row height — keeps Prompt collapsed; double-click to view full prompt
        shots.row_dimensions[row_num].height = 36

    # Column widths
    shots.column_dimensions["A"].width = 32     # Shot ID
    shots.column_dimensions["B"].width = 5      # L#
    shots.column_dimensions["C"].width = 45     # Description
    shots.column_dimensions["D"].width = 36     # References
    shots.column_dimensions["E"].width = 8      # Aspect
    shots.column_dimensions["F"].width = 14     # Style
    shots.column_dimensions["G"].width = 50     # Prompt — narrow → text clips
    shots.column_dimensions["H"].width = 10     # Status
    shots.column_dimensions["I"].width = 32     # v01
    shots.column_dimensions["J"].width = 32     # v02
    shots.column_dimensions["K"].width = 40     # Notes

    shots.freeze_panes = "A2"

    wb.save(output_path)

    stats = {
        "total": len(shots_data),
        "good": sum(1 for s in shots_data if s["status"] == "✓"),
        "partial": sum(1 for s in shots_data if s["status"] == "Partial"),
        "failed": sum(1 for s in shots_data if s["status"] == "✗"),
        "pending": sum(1 for s in shots_data if s["status"] == "pending"),
        "skipped": sum(1 for s in shots_data if s["status"] == "skipped"),
    }
    return stats


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("Usage: build_xlsx.py <config.json> <output.xlsx>")
    config_path, output_path = sys.argv[1], sys.argv[2]
    with open(config_path) as f:
        config = json.load(f)
    stats = build(config, output_path)
    print(f"✓ Wrote {output_path}")
    print(f"  {stats['total']} rows: {stats['good']} ✓ | {stats['partial']} partial | {stats['failed']} ✗ | {stats['pending']} pending | {stats['skipped']} skipped")
